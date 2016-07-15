
import numpy

from openpyxl import Workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill


from variation import GT_FIELD, CHROM_FIELD, POS_FIELD, REF_FIELD, ALT_FIELD, \
    MISSING_INT


COLORS = {'red': 'FFFF0000', 'green': 'FF00FF00', 'yellow': 'FFFFFF00',
          'blue': 'FF0000FF', 'purple': 'FFFF00FF', 'light_blue': 'FF00FFFF',
          'grey': 'FF888888'}

COLORS = {color: PatternFill(start_color=value, end_color=value,
                             fill_type='solid')
          for color, value in COLORS.items()}


GT_COLORS = {(0, 0): COLORS['red'], (1, 1): COLORS['green'],
             (0, 1): COLORS['yellow'], (0, 1): COLORS['yellow'],
             (2, 2): COLORS['blue'],
             (0, 2): COLORS['purple'], (2, 0): COLORS['purple'],
             (1, 2): COLORS['light_blue'], (2, 1): COLORS['light_blue']}


def _indi_snp_gt_to_str(snp_indi_gt):
    try:
        color_fill = GT_COLORS[tuple(snp_indi_gt)]
    except KeyError:
        color_fill = None
    print('snp_indi_gt', snp_indi_gt)
    if MISSING_INT in snp_indi_gt:
        color_fill = COLORS['grey']
    str_gt = '|'.join(map(str, snp_indi_gt))
    return str_gt, color_fill


def _create_allele_str(ref, alt):
    alleles = [ref]
    alleles.extend(alt)
    alleles = ','.join(map(str, alleles))
    return alleles


def _create_cell(worksheet, value):

    color_fill = None
    if issubclass(type(value), numpy.integer):
        value = int(value)
    elif isinstance(value, numpy.bytes_):
        value = str(value)
    elif hasattr(value, 'shape') or isinstance(value, (tuple, list)):
        value, color_fill = _indi_snp_gt_to_str(value)

    cell = WriteOnlyCell(worksheet, value=value)

    if color_fill is not None:
        cell.fill = color_fill
    return cell


def write_excel(variations, out_fhand):

    wbook = Workbook(write_only=True)
    sheet = wbook.create_sheet()

    samples = variations.samples

    chroms = variations[CHROM_FIELD] if CHROM_FIELD in variations else None
    poss = variations[POS_FIELD] if POS_FIELD in variations else None
    refs = variations[REF_FIELD] if REF_FIELD in variations else None
    alts = variations[ALT_FIELD] if ALT_FIELD in variations else None
    gts = variations[GT_FIELD]

    if samples:
        row = []
        if chroms is None:
            row.append(None)
        else:
            row.append('CHROM')
        if poss is None:
            row.append(None)
        else:
            row.append('POS')
        if refs is None:
            row.append(None)
        else:
            row.append('Alleles')

        row.extend([str(sample) for sample in samples])
        sheet.append(row)

    for snp_idx in range(gts.shape[0]):
        row = []
        chrom = None if chroms is None else chroms[snp_idx]
        row.append(chrom)
        pos = None if poss is None else poss[snp_idx]
        row.append(pos)

        if refs is None:
            alleles = None
        else:
            ref = refs[snp_idx]
            alt = alts[snp_idx]
            alleles = _create_allele_str(ref, alt)
        row.append(alleles)

        snp_gts = gts[snp_idx, :, :]
        row.extend(snp_gts)

        cells = [_create_cell(sheet, val) for val in row]
        sheet.append(cells)
    wbook.save(out_fhand)
    out_fhand.flush()
